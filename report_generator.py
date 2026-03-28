import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl

from config import SHEET_NAMES, HEADER_COLOR


def generate_report(df: pd.DataFrame, output_filename: str):
    """Генерирует красивый Excel-отчёт и график"""

    # Создаём сводные таблицы
    by_manager = df.groupby('Менеджер').agg({
        'Сумма_продажи': 'sum',
        'Прибыль': 'sum',
        'Количество': 'sum'
    }).round(0)

    by_category = df.groupby('Категория').agg({
        'Сумма_продажи': 'sum',
        'Прибыль': 'sum',
        'Количество': 'sum'
    }).round(0)

    top_goods = df.groupby('Товар')['Сумма_продажи'].sum().nlargest(10)

    # Сохраняем Excel с форматированием
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Лист 1: Общие данные
        df.to_excel(writer, sheet_name=SHEET_NAMES['main'], index=False)
        format_worksheet(writer.sheets[SHEET_NAMES['main']], df)

        # Лист 2: По менеджерам
        by_manager.to_excel(writer, sheet_name=SHEET_NAMES['managers'])
        format_worksheet(writer.sheets[SHEET_NAMES['managers']], by_manager, is_pivot=True)

        # Лист 3: По категориям
        by_category.to_excel(writer, sheet_name=SHEET_NAMES['categories'])
        format_worksheet(writer.sheets[SHEET_NAMES['categories']], by_category, is_pivot=True)

        # Лист 4: Топ-10 товаров
        top_goods.to_excel(writer, sheet_name=SHEET_NAMES['top'])
        format_worksheet(writer.sheets[SHEET_NAMES['top']], top_goods, is_pivot=True)

    print(f"Отчёт сохранён: {output_filename}")

    # Генерация графика
    create_sales_chart(by_category)


def format_worksheet(ws, df, is_pivot=False):
    """Применяет красивое форматирование к листу"""
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Форматирование шапки
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Автоширина колонок
    for idx, col in enumerate(ws.columns):
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 3, 40)
        ws.column_dimensions[column].width = adjusted_width

    # Денежный формат для числовых столбцов
    if not is_pivot:
        for row in ws.iter_rows(min_row=2):
            for cell in row[6:]:  # начиная примерно с 7-го столбца (суммы)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0 ₽'


def create_sales_chart(by_category: pd.DataFrame):
    """Создаёт график продаж по категориям"""
    plt.figure(figsize=(10, 6))

    sales = by_category['Сумма_продажи'].sort_values(ascending=False)

    ax = sales.plot(kind='bar', color='skyblue', edgecolor='black')
    plt.title('Сумма продаж по категориям товаров', fontsize=14, pad=20)
    plt.ylabel('Сумма продаж (руб.)', fontsize=12)
    plt.xlabel('Категория', fontsize=12)
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', alpha=0.3)

    # Подписи значений над столбцами
    for i, v in enumerate(sales):
        ax.text(i, v * 1.02, f'{int(v):,}', ha='center', fontsize=10)

    plt.tight_layout()
    plt.savefig('sales_by_category.png', dpi=200, bbox_inches='tight')
    plt.close()

    print("График сохранён: sales_by_category.png")